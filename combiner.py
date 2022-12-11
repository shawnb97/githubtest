from openpyxl import load_workbook 
import pandas as pd 
from pathlib import Path
from PyPDF2 import PdfFileMerger, PdfFileReader
import csv 
import os
import sys
from datetime import datetime
pdf_dir = Path(__file__).parent / "Paystubs"
pdf_dirt = Path(__file__).parent / "Timesheets"
pdf_output_dir = Path(__file__).parent / "OUTPUT"
pdf_output_dir.mkdir(parents = True, exist_ok=True)
Timesheets = list(pdf_dirt.glob("*.pdf"))
Paystubs= list(pdf_dir.glob("*.pdf"))
# file.name is the name of the file
#line[1] is empid from the excelsheet
#emplid is the file name 
#Date = set([file.name[0:10] for file in Paystubs])
#le in Paystubs:
#    empid = file.name[0:5]
#    if empid  == '11772':
#        print(file.name[6:16])
#wb = Path("C:\\Users\\jessica\\AppData\\Local\\Programs\\BHLABOR.xlsm")
#print(wb.read_text())
with open('BHLABOR.csv', 'r', encoding = "UTF-8") as f:
    csvr = csv.reader(f)
    for line in csvr:
 #       print(line[1])
 #       print('this is the person that is getting searched')
 #       if len(a) >
        shitlist= []
  #      a = len(shitlist)

            
        for file in Paystubs:
            try:
                df = datetime.strptime(line[5],'%m/%d/%Y').date()
            except ValueError:

                datefile=datetime.strptime(file.name[6:16],'%m.%d.%Y').date()
         
            emplid = file.name[0:5]
  
#            print(emplid)
            #this is the name id from the file 
            fullemplid = file.name
     
            if line[1] == emplid and df == datefile:# and dateexl ==datefile:#file.name[0:5]+file.name[6:16] xxxxx05.05.2022
                               #line[1]+line[5] xxxxx5/5/2022
                shitlist.append(fullemplid)

        for fileT in Timesheets:

            try:
                dp = datetime.strptime(line[5],'%m/%d/%Y').date()
                
            except ValueError:
                  datefileT = datetime.strptime(fileT.name.split('_')[1],'%m.%d.%Y').date()
                  
            emplidT = fileT.name[0:5]
#            print(emplid)
            #this is the name id from the file 

            fullemplidT = fileT.name
            if line[1] == emplidT and dp == datefileT:
                shitlist.append(fullemplidT)
            
        merger = PdfFileMerger()
        a = shitlist
        if len(a) >0:
            print(a)
            for item in a:
                item = item.replace("PDF", "pdf")
                if item[-5:-4] == 't':
                    main = "/Users/jessica/AppData/Local/Programs/Timesheets/"

                else:
                    main = "/Users/jessica/AppData/Local/Programs/Paystubs/"
                string = main + item
                merger.append(string)
            merger.write('./OUTPUT/'+item[0:6]+ line[2])

#                print(fullemplid)#datetime(line[5],"%m/%d/%Y"))
      


#        if line[1]= line[5]

  


#the key includes empoyee and data, EMPly id comes first 5 char nad date it char [6:16]


#wb_obj = openpyxl.load_workbook("C:\\Users\\jessica\\AppData\\Local\\Programs\\BHLABOR.xlsm")
#sheet_names = wb_obj.sheet_names()
#sheetz = wb_obj.sheet_by_name(sheet_names[0])
#cellz = sheetz.cell(row = 6, column =1)


#df = pd.read_excel('BHLABOR.xlsm','Employee PPE Summary',usecols=[1,4])


# the speed on thus one is ok 
#wb = load_workbook('BHLABOR.xlsx')
#ws = wb.active
#ws.title = "Employee PPE Summary"
#for row in range(6,20):
#    char = chr(74 + row)
#    char = chr(70)
#    print(ws[char +str(row)])
# the order we want is timesheet paystud time sheet pat stub in alternating order 
